import smtplib
from email.message import EmailMessage
from openpyxl import load_workbook

SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 587

sender = 'skdpa391@gmail.com'
password = 'pjhm srpc edwc qbjd'

load_wb = load_workbook('email_list.xlsx')
load_ws = load_wb['Sheet1']

for row in load_ws.iter_rows(min_row=1, values_only=True):
    recipient = row[0]  # 첫 번째 열의 값
    
    # 1. recipient가 None이 아니고, 빈 문자열이 아닐 때만 실행
    if recipient:
        msg = EmailMessage()
        msg['Subject'] = '휴강 공지'
        msg['From'] = sender    
        msg.set_content('오늘 수업은 휴강입니다.')
        msg['To'] = recipient

    s = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    s.ehlo()
    s.starttls()
    s.login(sender, password)
    s.send_message(msg)
    s.quit()